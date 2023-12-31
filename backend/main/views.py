from collections import defaultdict
from datetime import datetime, timedelta

from django.conf import settings
from django.http import JsonResponse, Http404, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
import json
from django.db import transaction
from django.db.models import Q, Sum, ExpressionWrapper, F, Max, Min, Max, Subquery, OuterRef, Count, Prefetch
from django.db.models.fields import DurationField, FloatField
from main.constants import ReportStatusCode
from main.utils import convert_string_to_date, get_chamber_status
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import mixins, viewsets, status
from rest_framework.exceptions import NotFound
from rest_framework.decorators import action, permission_classes, api_view
from rest_framework.response import Response
from main.permissions import HasAPIKeyOrIsAuthenticated

from main.filters import LotFilterSet, StatusReportFilterSet, NotificationFilterSet, LotDataFilterSet
from main.models import Lot, LotData, StatusReport, Notification
from main.pagination import CustomPageNumberPagination
from main.serializers import LotSerializer, LotDataSerializer, StatusReportSerializer, StatusReportListSerializer, LotExcelSerializer, LotDataExcelSerializer, NotificationSerializer
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

@require_GET
def get_all_lots(request):
    lots = Lot.objects.all().order_by("-start_time")
    data = [
        {"id": lot.id, "time": lot.start_time, "species": lot.species} for lot in lots
    ]
    return JsonResponse(data, safe=False)


@require_GET
def get_lots_in_time_period(request):
    start_time = request.GET.get("start_time")
    end_time = request.GET.get("end_time")
    lots = Lot.objects.filter(start_time__range=[start_time, end_time]).order_by(
        "-start_time"
    )
    data = [
        {"id": lot.id, "time": lot.start_time, "species": lot.species} for lot in lots
    ]
    return JsonResponse(data, safe=False)


@require_GET
def get_lots_by_species(request):
    species_list = request.GET.getlist("species")
    lots = Lot.objects.filter(species__in=species_list).order_by("-start_time")
    data = [
        {"id": lot.id, "time": lot.start_time, "species": lot.species} for lot in lots
    ]
    return JsonResponse(data, safe=False)


@require_GET
def get_lots_by_chamber(request):
    chamber = request.GET.get("chamber")
    lots = Lot.objects.filter(chamber=chamber).order_by("-start_time")
    data = [
        {"id": lot.id, "time": lot.start_time, "species": lot.species} for lot in lots
    ]
    return JsonResponse(data, safe=False)


@require_GET
def get_lot_details(request, lot_id):
    try:
        lot = Lot.objects.get(id=lot_id)
        data = {
            "id": lot.id,
            "company": lot.company.name,
            "chamber": lot.chamber,
            "start_time": lot.start_time,
            "complete_time": lot.complete_time,
            "program_name": lot.program_name,
            "total_commands": lot.total_commands,
            "species": lot.species,
            "quantity": lot.quantity,
            "details": lot.details,
        }
        return JsonResponse(data)
    except Lot.DoesNotExist:
        return JsonResponse({"error": "Lot not found"}, status=404)


@require_GET
def get_lotdata_by_lot(request, lot_id):
    try:
        lotdata = LotData.objects.filter(lot_id=lot_id).order_by("time")
        data = []
        for ld in lotdata:
            item = {
                "time": ld.time,
                "command_name": ld.command_name,
                "wbt1": ld.wbt1,
                "wbt2": ld.wbt2,
                "dbt1": ld.dbt1,
                "dbt2": ld.dbt2,
                "rh": ld.rh,
                "mc1": ld.mc1,
                "mc2": ld.mc2,
                "mc3": ld.mc3,
                "mc4": ld.mc4,
                "mc5": ld.mc5,
                "mc6": ld.mc6,
                "mc7": ld.mc7,
                "mc8": ld.mc8,
                "amc": ld.amc,
                "wood_temp1": ld.wood_temp1,
                "wood_temp2": ld.wood_temp2,
                "flaps": ld.flaps,
                "heat": ld.heat,
                "spray": ld.spray,
                "fan_cw": ld.fan_cw,
                "fan_ccw": ld.fan_ccw,
                "reserved": ld.reserved,
                "details": ld.details,
            }
            data.append(item)
        return JsonResponse(data, safe=False)
    except LotData.DoesNotExist:
        return JsonResponse({"error": "LotData not found"}, status=404)


"""
To developer: In production, all these POST endpoints will be used by the drying kilns computers
to post new data into the database.
As you can see, I have not implemented authorization and https, so please implement.
Please also provide sample Python code to post data using these endpoints.
"""


@csrf_exempt
@require_POST
def create_lot(request):
    try:
        data = json.loads(request.body)
        lot = Lot.objects.create(
            id=data["id"],
            company_id=data["company_id"],
            chamber=data["chamber"],
            start_time=data["start_time"],
            complete_time=data.get("complete_time", None),
            program_name=data.get("program_name", None),
            total_commands=data.get("total_commands", None),
            species=data.get("species", "none"),
            quantity=data.get("quantity", 0),
            details=data.get("details", None),
            duration=data.get("duration", None),
        )
        return JsonResponse({"success": True, "lot_id": lot.id}, status=201)
    except (KeyError, ValueError):
        return JsonResponse({"error": "Invalid data provided"}, status=400)


@csrf_exempt
@require_POST
def create_lotdata(request):
    try:
        data = json.loads(request.body)
        lotdata = LotData.objects.create(
            lot_id_id=data["lot_id"],
            time=data["time"],
            command_name=data["command_name"],
            wbt1=data["wbt1"],
            wbt2=data.get("wbt2", None),
            dbt1=data["dbt1"],
            dbt2=data.get("dbt2", None),
            rh=data.get("rh", None),
            mc1=data["mc1"],
            mc2=data["mc2"],
            mc3=data["mc3"],
            mc4=data["mc4"],
            mc5=data.get("mc5", None),
            mc6=data.get("mc6", None),
            mc7=data.get("mc7", None),
            mc8=data.get("mc8", None),
            amc=data["amc"],
            wood_temp1=data.get("wood_temp1", None),
            wood_temp2=data.get("wood_temp2", None),
            flaps=data.get("flaps", None),
            heat=data.get("heat", None),
            spray=data.get("spray", None),
            fan_cw=data.get("fan_cw", None),
            fan_ccw=data.get("fan_ccw", None),
            reserved=data.get("reserved", None),
            details=data.get("details", None),
        )
        return JsonResponse({"success": True, "lotdata_id": lotdata.id}, status=201)
    except (KeyError, ValueError):
        return JsonResponse({"error": "Invalid data provided"}, status=400)


"""
To developer:
I need 2 additional endpoints. 1 is for the kilns computer to report status (StatusReport model)
and the 2 one is to register Lot.complete_time of the current lot with value
once the drying lot is completed at the kiln.
"""


class LotViewSet(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.CreateModelMixin, mixins.RetrieveModelMixin, mixins.DestroyModelMixin):
    permission_classes = (HasAPIKeyOrIsAuthenticated,)
    filter_backends = (DjangoFilterBackend, )
    filterset_class = LotFilterSet
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        if self.request.user.is_superuser:
            qs = Lot.objects.all().order_by("-start_time")
        else:
            qs = Lot.objects.filter(company=self.request.user.company).order_by("-start_time")
        return qs
    
    def get_object(self):
        obj = super().get_object()
        if not self.request.user.is_superuser and self.request.user.company.id != obj.company.id:
            raise NotFound()
        return obj

    def get_serializer_class(self):
        return LotSerializer
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if self.request.query_params.get('get_all', 'False').lower() == 'true':
            serializer = self.get_serializer(instance=queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        paginated_queryset = self.paginate_queryset(queryset)
        serializer = self.get_serializer(instance=paginated_queryset, many=True)
        data = self.get_paginated_response(serializer.data)

        lot_summary = queryset.aggregate(
            total_quantity=Sum('quantity'),
            total_time_in_period=ExpressionWrapper(
                Max('complete_time') - Min('start_time'),
                output_field=DurationField()
            ),
            total_operation_time=Sum('duration'),
            occupancy_ratio=ExpressionWrapper(
                (F('total_operation_time') / F('total_time_in_period')) * 100,
                output_field=FloatField()
            )
        )

        response = {
            **data.data,
            "lot_summary": lot_summary
        }

        return Response(response, status=status.HTTP_200_OK)
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        if not request.user.is_superuser and request.user.company.id != request.data.get("company"):
            raise NotFound("Not found company.")
        
        # close old lots
        if request.user.is_superuser:
            uncomplete_lots = Lot.objects.filter(chamber=request.data.get("chamber"), complete_time__isnull=True)
        else:
            uncomplete_lots = Lot.objects.filter(chamber=request.data.get("chamber"), company=request.user.company, complete_time__isnull=True)
        
        uncomplete_lots.update(complete_time=datetime.now().replace(microsecond=0))

        return super().create(request, *args, **kwargs)
    
    @swagger_auto_schema(
        responses={200: ""}, )
    @action(
        detail=True, methods=['get'], url_name='lot_data', url_path='lot-data', filterset_class=LotDataFilterSet
    )
    def lot_lotdata(self, request, pk=None):
        if self.request.user.is_superuser:
            lot_data = LotData.objects.filter(lot_id=pk).order_by("-time")
        else:
            lot_data = LotData.objects.filter(lot_id=pk, lot__company=self.request.user.company).order_by("-time")

        if self.request.query_params.get('get_all', 'False').lower() == 'true':
            return Response(LotDataSerializer(instance=lot_data, many=True).data, status=status.HTTP_200_OK)

        lot_data = self.filter_queryset(lot_data)

        lot_data_summary = lot_data.aggregate(
            total_time=ExpressionWrapper(
                Max('time') - Min('time'),
                output_field=DurationField()
            ),
        )

        data = self.get_paginated_response(LotDataSerializer(instance=self.paginate_queryset(lot_data), many=True).data)
        
        response = {
            **data.data,
            "lot_data_summary": lot_data_summary
        }
        
        return Response(response, status=status.HTTP_200_OK)

    @swagger_auto_schema(responses={200: LotSerializer}, )
    @action(
        detail=False, methods=['get'], url_name='ongoing_lot', url_path='ongoing-lot'
    )
    def ongoing_lot(self, request):
        three_months_ago = datetime.now() - timedelta(days=settings.DEFAULT_DAYS)
        ongoing_lot = self.filter_queryset(Lot.objects.filter(
            complete_time__isnull=True, start_time__gte=three_months_ago
        ))
        if self.request.query_params.get('get_all', 'False').lower() == 'true':
            return Response(self.get_serializer(instance=ongoing_lot, many=True).data, status=status.HTTP_200_OK)
        ongoing_lot = self.paginate_queryset(ongoing_lot)
        data = self.get_paginated_response(self.get_serializer(instance=ongoing_lot, many=True).data)
        return data

    @swagger_auto_schema(responses={200: LotSerializer}, )
    @action(
        detail=False, methods=['get'], url_name='historical_lot', url_path='historical-lot'
    )
    def historical_lot(self, request):
        three_months_ago = datetime.now() - timedelta(days=settings.DEFAULT_DAYS)
        historical_lot = self.filter_queryset(Lot.objects.filter(
            Q(complete_time__isnull=False) | Q(start_time__lt=three_months_ago)
        ))
        if self.request.query_params.get('get_all', 'False').lower() == 'true':
            return Response(self.get_serializer(instance=historical_lot, many=True).data, status=status.HTTP_200_OK)
        historical_lot = self.paginate_queryset(historical_lot)
        data = self.get_paginated_response(self.get_serializer(instance=historical_lot, many=True).data)
        return data

    @swagger_auto_schema(responses={200: LotSerializer}, )
    @action(
        detail=True, methods=['get'], url_name='mark_lot_completed', url_path='mark-lot-completed'
    )
    def mark_lot_completed(self, request, pk=None):
        try:
            if request.user.is_superuser:
                completed_lot = Lot.objects.get(id=pk)
            else:
                completed_lot = Lot.objects.get(id=pk, company=request.user.company)
        
            completed_lot.complete_time = datetime.now().replace(microsecond=0)
            completed_lot.save()
            return Response(self.get_serializer(instance=completed_lot).data, status=status.HTTP_200_OK)
        except Lot.DoesNotExist:
            raise Http404

    @swagger_auto_schema(responses={200: ""})
    @action(
        detail=False, methods=['get'], url_name='download_lot_excel', url_path='lot-excel', pagination_class=None
    )
    def download_lot_excel(self, request):
        queryset =  self.filter_queryset(self.get_queryset())
        data = LotExcelSerializer(instance=queryset, many=True).data

        file_name = f'Lot'
        sheet_name = f'Lot'
        headers = ['Chamber', 'Lot ID', 'Program', 'Commands', 'Species', 'Quantity', 'Start Time', 'Complete Time', 'Ellapsed']
        
        response = export_excel(file_name, sheet_name, headers, data)
    
        return response
    
    @swagger_auto_schema(
        responses={200: ""}, )
    @action(
        detail=True, methods=['get'], url_name='download_lot_data_excel', url_path='lot-data-excel', filter_backends=[], pagination_class=None
    )
    def download_lot_data_excel(self, request, pk=None):
        if self.request.user.is_superuser:
            queryset = LotData.objects.filter(lot_id=pk).order_by("-time")
        else:
            queryset = LotData.objects.filter(lot_id=pk, lot__company=self.request.user.company).order_by("-time")
        
        data = LotDataExcelSerializer(instance=queryset, many=True).data

        file_name = f'[{pk}]LotData'
        sheet_name = f'{pk}_LotData'
        headers = ['ID', 'Time', 'Command', 'AMC', 'RH', 'DBT1', 'DBT2', 'Target DBT', 'WBT1', 'WBT2', 'Target WBT', 'MC1', 'MC2', 'MC3', 'MC4', 'MC5', 'MC6', 'MC7', 'MC8', 'Wood Temp 1', 'Wood Temp 2', 'Flaps', 'Heat', 'Spray', 'Fan CW', 'Fan CCW', 'Reserved', 'Details']
        response = export_excel(file_name, sheet_name, headers, data)
    
        return response


class LotDataViewSet(viewsets.GenericViewSet, mixins.CreateModelMixin):
    permission_classes = (HasAPIKeyOrIsAuthenticated,)
    serializer_class = LotDataSerializer

    def create(self, request, *args, **kwargs):
        try:
            if request.user.is_superuser:
                Lot.objects.get(id=request.data.get("lot"))
            else:
                Lot.objects.get(id=request.data.get("lot"), company=request.user.company)
        except Lot.DoesNotExist:
            raise NotFound("Not found lot.")

        return super().create(request, *args, **kwargs)


class StatusReportViewSet(viewsets.GenericViewSet, mixins.CreateModelMixin, mixins.ListModelMixin):
    permission_classes = (HasAPIKeyOrIsAuthenticated,)
    filter_backends = (DjangoFilterBackend, )
    filterset_class = StatusReportFilterSet
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        if self.request.user.is_superuser:
            queryset = StatusReport.objects.all()
        else:
            queryset = StatusReport.objects.filter(company=self.request.user.company)

        # Subquery to find the latest report for each chamber
        latest_report_subquery = (
            queryset
            .filter(chamber=OuterRef('chamber'))
            .order_by('-server_time')
            .values('id')[:1]
        )

        # Query to retrieve the latest StatusReport for each distinct chamber
        queryset = queryset.annotate(
            latest_id=Subquery(latest_report_subquery)
        ).filter(id=F('latest_id'))

        queryset = queryset.prefetch_related(Prefetch(
            "lot", queryset=Lot.objects.prefetch_related(Prefetch("lot_data", queryset=LotData.objects.order_by('-time')))
        )).order_by('chamber')
        
        return queryset

    def get_serializer_class(self):
        if self.action == 'create':
            return StatusReportSerializer
        return StatusReportListSerializer
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        if not self.request.user.is_superuser and request.user.company.id != request.data.get("company"):
            raise NotFound("Not found company.")
        
        try:
            if request.user.is_superuser:
                Lot.objects.get(id=request.data.get("lot"))
            else:
                Lot.objects.get(id=request.data.get("lot"), company=request.data.get("company"))
        except Lot.DoesNotExist:
            raise NotFound("Not found lot.")

        # Remove old reports of chamber
        StatusReport.objects.filter(
            company=request.data.get("company"),
            chamber=request.data.get("chamber")
        ).delete()

        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(responses={200: ''}, )
    @action(
        detail=False, methods=['get'], url_name='chamber_latest_status', url_path='chamber-latest-status'
    )
    def chamber_latest_status(self, request):
        queryset = self.get_queryset()

        if self.request.query_params.get('get_all', 'False').lower() == 'true':
            serializer = self.get_serializer(instance=queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        filter_queryset = self.filter_queryset(queryset)
        paginated_queryset = self.paginate_queryset(filter_queryset)
        
        serializer = self.get_serializer(instance=paginated_queryset, many=True)
        data = self.get_paginated_response(serializer.data)

        chamber_summary = queryset.aggregate(
            total_idle_chambers=Count('chamber', filter=Q(status_code=ReportStatusCode.IDLE)),
            total_operating_chambers=Count('chamber', filter=Q(status_code=ReportStatusCode.OPERATING)),
        )

        response = {
            **data.data,
            "chamber_summary": chamber_summary
        }

        return Response(response, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        responses={200: ""}, )
    @action(
        detail=False, methods=['get'], url_name='download_status_report_excel', url_path='excel', filter_backends=[], pagination_class=None
    )
    def download_status_report_excel(self, request, pk=None):
        queryset =  self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(instance=queryset, many=True)

        data = []
        for item in serializer.data:
            lot = item.get('lot') or {}
            lot_data = item.get('latest_lot_data') or {}

            data.append({
                'chamber': item.get('chamber'),
                'status': get_chamber_status(item.get('status_code')),
                'last_complete': item.get('last_complete'),
                'since_last_complete': item.get('since_last_complete'),
                'last_report': item.get('last_report'),
                'since_last_report': item.get('since_last_report'),
                'lot_id': lot.get('id'),
                'lot_species': lot.get('species'),
                'lot_quantity': lot.get('quantity'),
                'latest_lot_data_amc': lot_data.get('amc'),
                'latest_lot_data_dbt': lot_data.get('dbt'),
                'latest_lot_data_wbt': lot_data.get('wbt'),
                'total_time': item.get('total_time')
            })

        file_name = f'StatusReport'
        sheet_name = f'StatusReport'
        headers = ['Chamber', 'Status', 'Last Complete', 'Since Last Complete', 'Last Report', 'Since Last Report', 'Lot ID', 'Species', 'Quantity', 'AMC', 'DBT', 'WBT', 'Total Time']
        
        response = export_excel(file_name, sheet_name, headers, data)
        return response


class NotificationViewSet(viewsets.GenericViewSet, mixins.CreateModelMixin, mixins.ListModelMixin):
    permission_classes = (HasAPIKeyOrIsAuthenticated,)
    filter_backends = (DjangoFilterBackend, )
    filterset_class = NotificationFilterSet
    pagination_class = CustomPageNumberPagination
    serializer_class = NotificationSerializer

    def get_queryset(self):
        if self.request.user.is_superuser:
            queryset = Notification.objects.all().order_by("-time", "-id")
        else:
            queryset = Notification.objects.filter(company=self.request.user.company).order_by("-time", "-id")
        
        return queryset

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        if not self.request.user.is_superuser:
            if request.user.company.id != request.data.get("company"):
                raise NotFound("Not found company.")
            
        # Check if there are at least 200 records (records per company)
        if Notification.objects.filter(company=request.data.get("company")).count() >= 200: 
            # Get time of the notification at index 200 (order by time desc)
            max_time = Notification.objects.values('time', 'id').order_by('-time', '-id')[199]

            # Delete records that have the time <= found time
            Notification.objects.filter(time__lte=max_time['time'], pk__lte=max_time['id']).delete()

        return super().create(request, *args, **kwargs)
    

@swagger_auto_schema(methods=['get'], manual_parameters=[
    openapi.Parameter('start', openapi.IN_QUERY, type=openapi.TYPE_STRING, format="date-time"),
    openapi.Parameter('end', openapi.IN_QUERY, type=openapi.TYPE_STRING, format="date-time")
])
@api_view(["GET"])
@permission_classes([HasAPIKeyOrIsAuthenticated, ])
def statistic(request):
    start = convert_string_to_date(request.query_params.get('start'))
    end = convert_string_to_date(request.query_params.get('end'))
    print(start, end, request.query_params)
    if not start and not end:
        return Response(
            {'message': 'start time and end time is required'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    start = datetime.combine(start, datetime.min.time())
    end = datetime.combine(end + timedelta(days=1), datetime.min.time())
    time_period = end - start
    
    data = {
        'total_wood_dried': [],
        'total_chamber_quantity_dried': [],
    }

    if request.user.is_superuser:
        queryset = Lot.objects.all()
    else:
        queryset = Lot.objects.filter(company=request.user.company)

    completed_lot_queryset = queryset.filter(
        complete_time__date__range=[start, end], complete_time__isnull=False
    )

    wood_dried = completed_lot_queryset.values('species').annotate(total_quantity=Sum('quantity'))
    for wood in wood_dried:
        data['total_wood_dried'].append({
            'x': wood.get('species'),
            'y': wood.get('total_quantity')
        })

    chambers_quantities = completed_lot_queryset.values('chamber').annotate(total_quantity=Sum('quantity'))
    for chamber_quantity in chambers_quantities:
        data['total_chamber_quantity_dried'].append({
            'x': chamber_quantity.get('chamber'),
            'y': chamber_quantity.get('total_quantity')
        })
    
    lot_queryset = queryset.filter(
        Q(start_time__lte=end) &
        (Q(complete_time__isnull=True) | Q(complete_time__gte=start))
    )

    operating_time_dict = defaultdict(timedelta)
    idle_time_dict = defaultdict(timedelta)

    lots = lot_queryset.values()
    for lot in lots:
        chamber_id = int(lot.get('chamber'))
        start_time = lot.get('start_time') if lot.get('start_time') > start else start
        complete_time = lot.get('complete_time') if lot.get('complete_time') is not None and lot.get('complete_time') < end else end
        if start_time < complete_time:
            operation_time = complete_time - start_time
            operating_time_dict[chamber_id] += operation_time
            idle_time_dict[chamber_id] = time_period - operating_time_dict[chamber_id]

            # fix for super user: idle_time of chamber_id can be < 0 because operating_time of chamber is value of multi-company
            if idle_time_dict[chamber_id].total_seconds() < 0:
                idle_time_dict[chamber_id] = timedelta()

    data.update({
        'operation_time': operating_time_dict,
        'idle_time': idle_time_dict,
    })
    
    return Response(data, status=status.HTTP_200_OK)


def export_excel(file_name, sheet_name, headers, data):
    time_now = datetime.now().strftime("%Y%m%d%H%M")
    excel_filename = f"{file_name}_{time_now}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
   
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"{sheet_name}"

    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    for row_num, row in enumerate(data, 2):
        for col_num, cell_value in enumerate(row.values(), 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

            # Calculate the length of the cell content and set the column width
            col_letter = get_column_letter(col_num)
            column_dimension = worksheet.column_dimensions[col_letter]
            cell_length = len(str(cell_value))
            if column_dimension.width is None or column_dimension.width < cell_length:
                column_dimension.width = cell_length

    workbook.save(response)
    return response